"""
Gestor Automatizado de Reclamaciones B2B — Home Impex
Versión 3.0: Núcleo Robusto de Clasificación Multi-Criterio, Cruce en Cascada (EAN + SKU) y Relleno Oficial Integral
"""

import streamlit as st
import pandas as pd
import io
import os
import re
import glob
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from copy import copy
from datetime import datetime
import pymupdf

# ─────────────────────────────────────────────
#  CONFIG — Palabras clave y constantes del dominio
# ─────────────────────────────────────────────
INDICADORES_FACTURA = [
    'BIZONYLATSZÁM', 'BIZONYLATSZAM', 'INVOICE NR', 'INVOICE NUMBER', 'INVOICE NO',
    'INVOICE', 'SZÁMLASZÁM', 'SZAMLASZAM', 'PRICE/ITEM', 'UNIT PRICE', 'TOTAL PRICE',
    'TOTAL AMOUNT', 'PURCHASE ORDER NUMBER', 'QUANTITY CONFIRMED', 'INVOICE QTY',
    'PO PRICE', 'EXP. CART', 'GROSS WEIGHT', 'DELIVERY NOTE NO', 'DELIVERY NOTE',
]

INDICADORES_EAN = ['EAN', 'GTIN', 'BARCODE', 'BAR CODE', 'EAN CODE', 'EAN - GIFT BOX']

INDICADORES_SKU = [
    'ITEM NO.', 'ITEM NO', 'ITEM NUMBER', 'CIKKSZÁM', 'CIKKSZAM', 'SKU',
    'SUPPLIER CONFIG SKU', 'REF', 'REF.', 'ITEMCODE', 'ZALANDO CONFIG SKU', 'ARTICLE NUMBER'
]

INDICADORES_QTY_QUEJA = [
    '# DEFECT', 'DEFECT ITEM', 'DEFECT COUNT', 'DEFECT QTY',
    'MANQUANTS', 'DIFFERENCE', 'MISSING', 'SHORTAGE',
    'REWORKED ITEM', 'REWORK COUNT', 'QUANTITY CONCERNED',
    'QUANTITY DEFECT', 'QTY DEFECT', 'QTY CONCERNED',
    'REWORKING', 'AFFECTED'
]

INDICADORES_REASON = [
    'DEFECT REASON', 'REASON', 'MOTIVE', 'COMMENT', 'RSC TASK', 'TASK',
    'PANASZ OKA', 'REASON FOR COMPLAINT'
]

INDICADORES_PLANTILLA = {
    'ref':           ['CIKKSZÁM', 'ITEM NUMBER', 'ARTICLE NUMBER'],
    'name':          ['MEGNEVEZÉS', 'NAME', 'PRODUCT NAME'],
    'delivery_date': ['KISZÁLLÍTÁS', 'DATE OF DELIVERY', 'DELIVERY DATE'],
    'qty_comprada':  ['VÁSÁROLT', 'PURCHASED PCS'],
    'qty_defect':    ['ÉRINTETT', 'QUANTITY CONCERNED', 'REWORKING', 'AFFECTED'],
    'precio':        ['NETTÓ ÁR', 'NET PRICE', 'PRICE/PCS', 'PRICE'],
    'invoice':       ['SZÁMLASZÁM', 'INVOICE NUMBER', 'INVOICE NO'],
    'reason':        ['PANASZ OKA', 'REASON FOR COMPLAINT', 'REASON'],
}


CARPETA_CATALOGO = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'catalogo')

REASON_OPCIONES = [
    "WORKMANSHIP", "MATERIAL", "BROKEN HANDLE", "BROKEN LID",
    "BROKEN LID HANDLE", "CHANGED COLOUR", "DAMAGED", "DAMAGED BOX",
    "DAMAGED COATING", "DEFORMED", "DENTED", "DOES NOT BOIL",
    "LEAKING", "LID DOES NOT FIT", "LOOSE HANDLE", "MISSING PIECES",
    "NOT TURNING ON", "NOT WARMING UP", "NOT WORKING",
    "NOT WORKING DISPLAY", "NOT WORKING ON INDUCTION",
    "RUSTY", "SCRATCHED COATING", "SHORT-CIRCUITED",
    "WORN PAINT", "WRONG ARTWORK", "RSC - POLYBAG",
    "RSC - EAN LABELING", "RSC - PACKAGING",
]

REQUEST_OPCIONES = [
    "Pay penalty", "Credit note", "Replacement", "Return and credit", "Scrap on site"
]

# ─────────────────────────────────────────────
#  HELPERS DE NORMALIZACIÓN
# ─────────────────────────────────────────────

def norm_sku(s) -> str:
    """Normaliza un SKU para comparaciones insensibles a caracteres especiales y espacios."""
    if not s or pd.isna(s):
        return ''
    return re.sub(r'[^A-Z0-9]', '', str(s).upper().strip())

def clean_ean(val) -> str:
    """Limpia y valida un código de barras EAN/GTIN."""
    if not val or pd.isna(val):
        return ''
    try:
        s = str(int(float(str(val).replace(' ', ''))))
        return s if re.match(r'^\d{8,14}$', s) else ''
    except (ValueError, TypeError):
        s_str = re.sub(r'\D', '', str(val))
        return s_str if re.match(r'^\d{8,14}$', s_str) else ''

def format_date(s) -> str:
    """Formatea cualquier fecha a formato limpio DD.MM.YYYY."""
    if not s or pd.isna(s):
        return ''
    s_str = str(s).strip()
    if ' ' in s_str and (':' in s_str or '-' in s_str):
        s_str = s_str.split(' ')[0]
    clean = re.sub(r'\s+', '', s_str)
    # DD.MM.YYYY
    m1 = re.match(r'^(\d{1,2})\.(\d{1,2})\.(\d{4})$', clean)
    if m1:
        d, mth, y = m1.groups()
        return f'{int(d):02d}.{int(mth):02d}.{y}'
    # YYYY-MM-DD o YYYY.MM.DD
    m2 = re.match(r'^(\d{4})[-\.](\d{1,2})[-\.](\d{1,2})', clean)
    if m2:
        y, mth, d = m2.groups()
        return f'{int(d):02d}.{int(mth):02d}.{y}'
    return clean

# ─────────────────────────────────────────────
#  CATÁLOGO MAESTRO — Carga e indexación dual
# ─────────────────────────────────────────────

@st.cache_data(show_spinner=False)
def cargar_catalogo_maestro(carpeta: str, override_bytes=None, override_name='') -> tuple:
    """
    Carga el catálogo maestro e indexa por EAN y por SKU normalizado.
    Prioridad: archivo subido manualmente > archivo más reciente en /catalogo/.
    """
    df = None
    info = ""

    if override_bytes:
        try:
            df = pd.read_excel(io.BytesIO(override_bytes))
            info = f"📂 Override: `{override_name}` ({len(df):,} filas)"
        except Exception as e:
            return {'by_ean': {}, 'by_sku': {}}, f"⚠️ Error catálogo manual: {e}"
    else:
        patron = os.path.join(carpeta, '*.xlsx')
        archivos = glob.glob(patron)
        if not archivos:
            patron_xls = os.path.join(carpeta, '*.xls')
            archivos = glob.glob(patron_xls)

        if not archivos:
            return {'by_ean': {}, 'by_sku': {}}, "⚠️ Sin catálogo en `/catalogo/`"

        archivos.sort(key=os.path.getmtime, reverse=True)
        mas_reciente = archivos[0]
        try:
            df = pd.read_excel(mas_reciente)
            nombre_f = os.path.basename(mas_reciente)
            fecha_mod = datetime.fromtimestamp(os.path.getmtime(mas_reciente)).strftime('%d/%m/%Y')
            info = f"📚 Catálogo: `{nombre_f}` ({fecha_mod}) · {len(df):,} productos"
        except Exception as e:
            return {'by_ean': {}, 'by_sku': {}}, f"⚠️ Error al leer catálogo: {e}"

    df.columns = [str(c).strip() for c in df.columns]

    col_ean = next((c for c in df.columns if any(e in c.upper() for e in ['EAN - GIFT BOX', 'EAN CODE', 'EAN', 'BARCODE'])), None)
    col_sku = next((c for c in df.columns if any(s in c.upper() for s in ['ITEM NO.', 'ITEM NUMBER', 'CIKKSZÁM', 'CIKKSZAM', 'SKU', 'REF'])), None)
    col_name = next((c for c in df.columns if 'PRODUCT NAME' == c.upper().strip()), None) or \
               next((c for c in df.columns if 'NAME' in c.upper()), None)
    col_detail = next((c for c in df.columns if 'DETAILED' in c.upper()), None)

    if not col_sku:
        return {'by_ean': {}, 'by_sku': {}}, '⚠️ Catálogo sin columna Item no.'

    catalogo_ean = {}
    catalogo_sku = {}

    for _, row in df.iterrows():
        try:
            sku_val = str(row[col_sku]).strip() if pd.notna(row[col_sku]) else ''
            name_val = str(row[col_name]).strip() if col_name and pd.notna(row[col_name]) else ''
            detail_val = str(row[col_detail]).strip() if col_detail and pd.notna(row[col_detail]) else ''
            
            data = {
                'sku':    sku_val,
                'name':   name_val or detail_val,
                'detail': detail_val,
            }
            
            if col_ean:
                ean = clean_ean(row.get(col_ean, ''))
                if ean:
                    catalogo_ean[ean] = data

            if sku_val:
                catalogo_sku[norm_sku(sku_val)] = data
        except Exception:
            continue

    return {'by_ean': catalogo_ean, 'by_sku': catalogo_sku}, info

# ─────────────────────────────────────────────
#  EXTRACCIÓN DE FACTURAS PDF
# ─────────────────────────────────────────────

def extraer_pdf_factura(file_bytes) -> pd.DataFrame:
    """
    Extrae datos de una factura PDF Home Impex / Zalando.
    Renderiza temporalmente cada página en RAM para inspección estructurada y
    extrae de forma limpia Invoice No., Delivery date, SKU, EAN, Cantidad y Precio.
    """
    RE_INVOICE = re.compile(r'Invoice No\.:\s*([A-Z]\d{2}/\d{5})', re.IGNORECASE)
    RE_DATE_EXTRACT = re.compile(
        r'(?:Invoice date/Sz[aá]mla kelte|Delivery date/Teljes[íi]t[eé]s d[aá]tuma|Kelte|Date)\s*[:\s]+(\d{1,2}\s*\.\s*\d{1,2}\s*\.\s*\d{4}|\d{4}[-\.\/]\d{1,2}[-\.\/]\d{1,2}|\d{1,2}[-\.\/]\d{1,2}[-\.\/]\d{4})',
        re.IGNORECASE
    )
    RE_ITEM = re.compile(
        r'^([A-Z]{2}[-/][\w/\-]+)\s+(\d{7,10})\s+(\d[\d\s]*)\s+db\s+(\d[\d\s]*,\d{2})\s+(\d[\d\s]*,\d{2})\s*$'
    )
    RE_EAN = re.compile(r'EAN:\s*(\d{8,14})')

    meta = {'invoice_no': '', 'delivery_date': ''}
    articulos = []

    try:
        doc = pymupdf.open(stream=file_bytes.read(), filetype='pdf')
    except Exception:
        return pd.DataFrame()

    # Si es un documento de abono (CREDIT NOTE), no extraer como factura de venta
    first_page_text = doc[0].get_text() if len(doc) > 0 else ''
    if 'CREDIT NOTE' in first_page_text.upper():
        doc.close()
        return pd.DataFrame()

    for page in doc:
        # Renderizado en memoria (144 DPI) para garantizar soporte de capas complejas
        mat = pymupdf.Matrix(2, 2)
        pix = page.get_pixmap(matrix=mat, colorspace=pymupdf.csGRAY)
        _img = pix.tobytes('png')

        text = page.get_text('text', sort=True)
        lines = [l.strip() for l in text.splitlines() if l.strip()]

        if not meta['invoice_no']:
            m_inv = RE_INVOICE.search(text)
            if m_inv:
                meta['invoice_no'] = m_inv.group(1).strip()

        if not meta['delivery_date']:
            m_date = RE_DATE_EXTRACT.search(text)
            if m_date:
                meta['delivery_date'] = format_date(m_date.group(1))

        i = 0
        while i < len(lines):
            m_item = RE_ITEM.match(lines[i])
            if m_item:
                sku   = m_item.group(1).strip()
                qty   = int(m_item.group(3).replace(' ', ''))
                price = float(m_item.group(4).replace(' ', '').replace(',', '.'))
                ean, name_en = '', ''

                if i + 1 < len(lines):
                    m_ean = RE_EAN.search(lines[i + 1])
                    if m_ean:
                        ean = m_ean.group(1)

                for j in range(i + 2, min(i + 5, len(lines))):
                    l = lines[j]
                    if RE_ITEM.match(l) or RE_EAN.search(l):
                        break
                    if len(l) >= 4:
                        name_en = (name_en + ' ' + l).strip() if name_en else l

                if sku or ean:
                    articulos.append({
                        'Bizonylatszám': meta['invoice_no'],
                        'Delivery date': meta['delivery_date'],
                        'Item. no.':     sku,
                        'EAN code':      ean,
                        'EAN_clean':     clean_ean(ean),
                        'Product name':  name_en,
                        'Quantity':      qty,
                        'Price/item':    price,
                        '_source':       'pdf',
                    })
                i += 2
                continue
            i += 1

    doc.close()
    return pd.DataFrame(articulos)

# ─────────────────────────────────────────────
#  LECTOR UNIVERSAL DE ARCHIVOS EXCEL / CSV / PDF
# ─────────────────────────────────────────────

def leer_hoja_excel(file_bytes, sheet_name: str, engine: str) -> pd.DataFrame | None:
    """Busca dinámicamente la fila de cabeceras en una hoja analizando las primeras 30 filas."""
    try:
        df_raw = pd.read_excel(file_bytes, sheet_name=sheet_name, header=None, nrows=30, engine=engine)
    except Exception:
        return None

    header_row = -1
    for i, row in df_raw.iterrows():
        row_str = ' '.join(str(v).upper() for v in row if pd.notna(v))
        has_ean = any(ind in row_str for ind in INDICADORES_EAN)
        has_sku = any(ind in row_str for ind in INDICADORES_SKU)
        has_qty = any(ind in row_str for ind in INDICADORES_QTY_QUEJA + ['QUANTITY', 'QTY', 'MENNYISÉG'])
        if (has_ean or has_sku) and has_qty:
            header_row = i
            break
        elif has_ean and has_sku:
            header_row = i
            break

    if header_row == -1:
        return None

    file_bytes.seek(0)
    try:
        df = pd.read_excel(file_bytes, sheet_name=sheet_name, header=header_row, engine=engine)
        df.columns = [str(c).strip() for c in df.columns]
        return df
    except Exception:
        return None

def leer_archivo(uploaded_file) -> tuple:
    """
    Clasifica y procesa cualquier archivo como 'factura', 'queja' o descarta.
    Soporta múltiples hojas, archivos con solo SKU, y facturas en formato tabla.
    """
    nombre = uploaded_file.name
    nombre_lower = nombre.lower()

    if 'complaint' in nombre_lower and 'form' in nombre_lower:
        return None, 'ignorar'

    file_bytes = io.BytesIO(uploaded_file.read())

    # ── 1. Archivo PDF ───────────────────────
    if nombre.endswith('.pdf'):
        df = extraer_pdf_factura(file_bytes)
        if df.empty:
            return None, 'pdf_sin_datos'
        return df, 'factura'

    # ── 2. Archivo CSV ───────────────────────
    if nombre.endswith('.csv'):
        try:
            df = pd.read_csv(file_bytes, sep=None, engine='python', encoding='utf-8')
            df.columns = [str(c).strip() for c in df.columns]
        except Exception as e:
            return None, f'error: {e}'
        
        cols_upper = [str(c).upper() for c in df.columns]
        es_factura = any(any(ind in c for ind in INDICADORES_FACTURA) for c in cols_upper)
        
        # Identificar EAN o SKU
        col_ean = next((c for c in df.columns if any(ind in str(c).upper() for ind in INDICADORES_EAN)), None)
        col_ref = next((c for c in df.columns if str(c).upper().strip() in ['REF.', 'REF', 'SUPPLIER CONFIG SKU', 'SUPPLIER SKU']), None)
        col_sku = col_ref or next((c for c in df.columns if any(ind in str(c).upper() for ind in ['ITEM. NO.', 'ITEM NO', 'CIKKSZÁM', 'CIKKSZAM', 'ITEMCODE', 'SKU', 'ARTICLE NUMBER'])), None) or \
                  next((c for c in df.columns if any(ind in str(c).upper() for ind in INDICADORES_SKU)), None)
        
        if not col_ean and not col_sku:
            return None, 'sin_ean'

        df['EAN_clean'] = df[col_ean].apply(clean_ean) if col_ean else ''
        if col_sku:
            df['Ref.'] = df[col_sku].astype(str)

        return df, 'factura' if es_factura else 'queja'

    # ── 3. Archivo Excel (.xlsx / .xls) ──────
    engine = 'openpyxl' if nombre.endswith('.xlsx') else 'xlrd'
    try:
        xl = pd.ExcelFile(file_bytes, engine=engine)
    except Exception as e:
        return None, f'error: {e}'

    # Heurística por nombre de archivo
    prioridad_factura = any(k in nombre_lower for k in ['invoice', 'szamla', 'számla', 'faktura'])
    prioridad_queja = any(k in nombre_lower for k in ['defect', 'manquant', 'queja', 'reclamacion', 'rsc', 'dnz'])

    hojas_procesadas = []

    for hoja in xl.sheet_names:
        file_bytes.seek(0)
        df_sheet = leer_hoja_excel(file_bytes, hoja, engine)
        if df_sheet is None or df_sheet.empty:
            continue

        cols_upper = [str(c).upper() for c in df_sheet.columns]
        col_str_all = ' '.join(cols_upper)

        col_ean = next((c for c in df_sheet.columns if any(ind in str(c).upper() for ind in INDICADORES_EAN)), None)
        
        # Prioridad de SKU: Ref. / Supplier Config SKU > Item no. / Cikkszám > Zalando SKU
        col_ref = next((c for c in df_sheet.columns if str(c).upper().strip() in ['REF.', 'REF', 'SUPPLIER CONFIG SKU', 'SUPPLIER SKU']), None)
        col_sku = col_ref or next((c for c in df_sheet.columns if any(ind in str(c).upper() for ind in ['ITEM. NO.', 'ITEM NO', 'CIKKSZÁM', 'CIKKSZAM', 'ITEMCODE', 'SKU', 'ARTICLE NUMBER'])), None) or \
                  next((c for c in df_sheet.columns if any(ind in str(c).upper() for ind in INDICADORES_SKU)), None)

        if not col_ean and not col_sku:
            continue

        df_sheet['EAN_clean'] = df_sheet[col_ean].apply(clean_ean) if col_ean else ''
        if col_sku:
            df_sheet['Ref.'] = df_sheet[col_sku].astype(str)

        # Determinar si es Factura o Queja
        tiene_defectos = any(any(ind in c for ind in INDICADORES_QTY_QUEJA) for c in cols_upper) or \
                         any(ind in col_str_all for ind in ['DEFECT REASON', 'DEFECT ITEM', 'DEFECT COUNT', 'RSC TASK', 'MANQUANTS'])

        tiene_factura = any(any(ind in c for ind in INDICADORES_FACTURA) for c in cols_upper)

        if prioridad_factura:
            tipo = 'factura'
        elif prioridad_queja and tiene_defectos:
            tipo = 'queja'
        elif 'PACKING LIST' in nombre_lower:
            tipo = 'factura'
        elif tiene_defectos:
            tipo = 'queja'
        elif tiene_factura:
            tipo = 'factura'
        else:
            col_r_test = next((c for c in df_sheet.columns if any(r in str(c).upper() for r in INDICADORES_REASON)), None)
            tipo = 'queja' if col_r_test else 'factura'


        # Formatear columnas según el tipo
        if tipo == 'factura':
            col_date = next(
                (c for c in df_sheet.columns if any(k in str(c).upper()
                 for k in ['KELTE', 'DATE', 'DELIVERY', 'TELJESÍTÉS', 'TELJESITES', 'ISSUE DATE'])), None
            )
            if col_date:
                df_sheet['Delivery date'] = df_sheet[col_date].apply(format_date)
            hojas_procesadas.append((df_sheet, 'factura'))

        else:
            # Es Queja: localizar cantidad afectada y motivo
            col_qty = None
            for c in df_sheet.columns:
                if any(p in str(c).upper() for p in INDICADORES_QTY_QUEJA):
                    sample = pd.to_numeric(df_sheet[c], errors='coerce')
                    if sample.notna().sum() > 0:
                        col_qty = c
                        break
            if not col_qty:
                for c in df_sheet.columns:
                    if any(p in str(c).upper() for p in ['QUANTITY', 'QTY', 'MENNYISÉG', '#']):
                        sample = pd.to_numeric(df_sheet[c], errors='coerce')
                        if sample.notna().sum() > 0:
                            col_qty = c
                            break

            df_sheet['Cantidad_Defecto'] = pd.to_numeric(df_sheet[col_qty], errors='coerce').abs() if col_qty else 1
            df_sheet = df_sheet[df_sheet['Cantidad_Defecto'] > 0].copy()

            col_r = next(
                (c for c in df_sheet.columns if any(r in str(c).upper() for r in INDICADORES_REASON)), None
            )
            df_sheet['_reason_detected'] = df_sheet[col_r].astype(str) if col_r else ''
            hojas_procesadas.append((df_sheet, 'queja'))

    if not hojas_procesadas:
        return None, 'sin_ean'

    # Si hay varias hojas (ej. hoja de quejas con más filas), seleccionar la de mayor cantidad de datos
    hojas_procesadas.sort(key=lambda x: len(x[0]), reverse=True)
    return hojas_procesadas[0]

def extract_invoice_fields(row):
    """Extrae de forma unificada y segura todos los campos contables de cualquier fila de factura o packing list."""
    inv = ''
    for col in ['Bizonylatszám', 'Bizonylatszam', 'Invoice nr.', 'Invoice nr', 'Invoice no.', 'Invoice no', 'Invoice No', 'Document no.', 'Delivery note no.', 'Számlaszám', 'Purchase Order Number']:
        if col in row and pd.notna(row[col]) and str(row[col]).strip() not in ('', 'nan', 'None'):
            inv = str(row[col]).strip()
            break

    date_val = ''
    for col in ['Delivery date', 'Invoice date', 'Kelte', 'Issue date', 'Date', 'Számla kelte', 'Teljesítés dátuma']:
        if col in row and pd.notna(row[col]) and str(row[col]).strip() not in ('', 'nan', 'None', 'NaT'):
            d_str = format_date(row[col])
            if re.search(r'\d{4}|\d{1,2}\.\d{1,2}', d_str) and not re.match(r'^\d+\.0$', d_str):
                date_val = d_str
                break

    qty_val = 0
    for col in ['Quantity', 'Qty', 'Total delivery', 'Mennyiség', 'Quantity Confirmed', 'Invoice Qty']:
        if col in row and pd.notna(row[col]):
            try:
                q = float(str(row[col]).replace(' ', ''))
                if q > 0:
                    qty_val = int(q)
                    break
            except Exception:
                pass

    price_val = None
    for col in ['Price/item', 'Unit price', 'Price', 'Nettó egységár', 'PO Price', 'Purchase Order Price before Discount']:
        if col in row and pd.notna(row[col]):
            try:
                p = float(str(row[col]).replace(' ', '').replace(',', '.'))
                if p > 0:
                    price_val = p
                    break
            except Exception:
                pass

    name_val = ''
    for col in ['Product name', 'Item name', 'Megnevezés', 'Detailed product name', 'Supplier Article Name', 'Name']:
        if col in row and pd.notna(row[col]) and str(row[col]).strip() not in ('', 'nan', 'None'):
            name_val = str(row[col]).replace('\n', ' ').strip()
            break

    sku_val = ''
    for col in ['Item. no.', 'Item no.', 'Itemcode', 'Item no', 'Cikkszám', 'Cikkszam', 'SKU', 'Ref.', 'Supplier Config SKU']:
        if col in row and pd.notna(row[col]) and str(row[col]).strip() not in ('', 'nan', 'None'):
            sku_val = str(row[col]).strip()
            break

    return inv, date_val, qty_val, price_val, name_val, sku_val

# ─────────────────────────────────────────────
#  MOTOR DE CRUCE EN CASCADA (EAN + SKU + HISTÓRICO)
# ─────────────────────────────────────────────

def cruzar_datos(df_quejas, df_facturas, catalogo, reason_default, fecha_manual_override='', factura_manual_override='', partner_request='Pay penalty', moneda='€') -> tuple:
    """
    Cruza cada artículo en cascada multi-nivel:
    1A. Match por EAN en facturas subidas -> recupera SKU, nombre, Nº factura, fecha, qty comprada, precio.
    1B. Match por SKU normalizado en facturas subidas -> recupera mismos datos contables.
    2A. Match por EAN en Catálogo Maestro -> recupera SKU y nombre. Fallback precio del reporte.
    2B. Match por SKU en Catálogo Maestro -> recupera nombre. Fallback precio del reporte.
    3. Sin datos -> Metadatos generales y precio del reporte.
    """
    resultados = []
    no_encontrados = 0

    cat_ean = catalogo.get('by_ean', {}) if isinstance(catalogo, dict) else {}
    cat_sku = catalogo.get('by_sku', {}) if isinstance(catalogo, dict) else {}

    # ── 1. Construir base de facturas indexada por EAN y por SKU ───────────────
    facturas_by_ean = {}
    facturas_by_sku = {}
    todas_fechas = []
    todas_facturas = []

    if df_facturas is not None and not df_facturas.empty:
        for _, frow in df_facturas.iterrows():
            inv_no, d_date, q_val, p_val, n_val, sku_raw = extract_invoice_fields(frow)
            ean_f = clean_ean(frow.get('EAN_clean', frow.get('EAN code', frow.get('EAN', ''))))
            sku_norm_f = norm_sku(sku_raw)

            if d_date and d_date not in ('nan', 'NaT', ''):
                todas_fechas.append(d_date)
            if inv_no and inv_no not in ('nan', '(not found)', ''):
                todas_facturas.append(inv_no)

            f_record = {
                'invoice_no':    inv_no,
                'delivery_date': d_date,
                'purchased_qty': q_val,
                'unit_price':    p_val,
                'sku':           sku_raw,
                'name':          n_val,
            }

            def _parse_date_sort(d_str):
                if not d_str: return ''
                clean = re.sub(r'\s+', '', str(d_str))
                m = re.match(r'^(\d{1,2})\.(\d{1,2})\.(\d{4})$', clean)
                if m:
                    d, mth, y = m.groups()
                    return f"{y}-{int(mth):02d}-{int(d):02d}"
                return str(d_str)

            def _update_record(target_dict, key, rec):
                if not key: return
                if key not in target_dict:
                    target_dict[key] = rec
                else:
                    existing = target_dict[key]
                    d_new = _parse_date_sort(rec['delivery_date'])
                    d_old = _parse_date_sort(existing['delivery_date'])
                    inv_new = rec['invoice_no']
                    inv_old = existing['invoice_no']

                    overwrite = False
                    if inv_new and not inv_old:
                        overwrite = True
                    elif inv_new and inv_old:
                        if d_new and d_old and d_new > d_old:
                            overwrite = True
                        elif inv_new > inv_old and (not d_old or d_new >= d_old):
                            overwrite = True
                        elif rec['purchased_qty'] > existing['purchased_qty'] and (not d_old or d_new >= d_old):
                            overwrite = True

                    if overwrite:
                        target_dict[key] = {
                            'invoice_no':    rec['invoice_no'] or existing['invoice_no'],
                            'delivery_date': rec['delivery_date'] or existing['delivery_date'],
                            'purchased_qty': rec['purchased_qty'] if rec['purchased_qty'] > 0 else existing['purchased_qty'],
                            'unit_price':    rec['unit_price'] if rec['unit_price'] is not None else existing['unit_price'],
                            'sku':           rec['sku'] or existing['sku'],
                            'name':          rec['name'] or existing['name'],
                        }
                    else:
                        if not existing['invoice_no'] and rec['invoice_no']:
                            existing['invoice_no'] = rec['invoice_no']
                        if not existing['delivery_date'] and rec['delivery_date']:
                            existing['delivery_date'] = rec['delivery_date']
                        if existing['purchased_qty'] <= 0 and rec['purchased_qty'] > 0:
                            existing['purchased_qty'] = rec['purchased_qty']
                        if existing['unit_price'] is None and rec['unit_price'] is not None:
                            existing['unit_price'] = rec['unit_price']
                        if not existing['name'] and rec['name']:
                            existing['name'] = rec['name']

            if ean_f:
                _update_record(facturas_by_ean, ean_f, f_record)
            if sku_norm_f:
                _update_record(facturas_by_sku, sku_norm_f, f_record)

    fechas_unicas = [f for f in todas_fechas if f and f not in ('nan', 'NaT', '')]
    facturas_unicas = [f for f in todas_facturas if f and f not in ('nan', '(not found)', '')]

    # Determinación inteligente de valores por defecto (batch fallback)
    if fecha_manual_override and fecha_manual_override.strip():
        fecha_default = format_date(fecha_manual_override.strip())
    elif fechas_unicas:
        fecha_default = fechas_unicas[-1]
    else:
        fecha_default = datetime.now().strftime('%d.%m.%Y')

    if factura_manual_override and factura_manual_override.strip():
        factura_default = factura_manual_override.strip()
    elif facturas_unicas:
        factura_default = facturas_unicas[-1]
    else:
        factura_default = ''


    # ── 2. Cruce fila a fila de quejas ────────────────────────────────────────
    for _, row in df_quejas.iterrows():
        ean = clean_ean(row.get('EAN_clean', row.get('EAN', '')))
        qty_def = int(pd.to_numeric(row.get('Cantidad_Defecto', 1), errors='coerce') or 1)
        reason_raw = str(row.get('_reason_detected', '')).strip()
        reason = reason_raw if reason_raw not in ('', 'nan') else reason_default

        ref_raw = str(row.get('Ref.', row.get('Réf. fourn.', row.get('Supplier Config SKU', row.get('Item. no.', row.get('CIKKSZÁM', '')))))).strip()
        if ref_raw in ('nan', '[Sin Ref]', ''):
            ref_raw = ''
        sku_norm_q = norm_sku(ref_raw)

        precio_queja = pd.to_numeric(
            row.get('Average Purchase Price', row.get('Reworking Costs', row.get('Prix achat € HT', None))),
            errors='coerce'
        )

        # ── NIVEL 1A: Match Factura por EAN
        if ean and ean in facturas_by_ean:
            f_match = facturas_by_ean[ean]
            c_data = cat_ean.get(ean) or cat_sku.get(sku_norm_q)
            sku_final = (c_data['sku'] if c_data else None) or (f_match['sku'] if f_match['sku'] and not f_match['sku'].startswith('ZZO') else None) or ref_raw
            nombre_final = (c_data['detail'] or c_data['name']) if c_data else (f_match['name'] or '')
            fecha_final = f_match['delivery_date'] or fecha_default
            factura_final = f_match['invoice_no'] or factura_default
            qty_cmp_final = f_match['purchased_qty'] if f_match['purchased_qty'] > 0 else qty_def
            precio_final = precio_queja if (pd.notna(precio_queja) and float(precio_queja) > 0) else f_match['unit_price']
            fuente = '✅ Factura (EAN)'

        # ── NIVEL 1B: Match Factura por SKU
        elif sku_norm_q and sku_norm_q in facturas_by_sku:
            f_match = facturas_by_sku[sku_norm_q]
            c_data = cat_sku.get(sku_norm_q) or cat_ean.get(ean)
            sku_final = (c_data['sku'] if c_data else None) or (f_match['sku'] if f_match['sku'] and not f_match['sku'].startswith('ZZO') else None) or ref_raw
            nombre_final = (c_data['detail'] or c_data['name']) if c_data else (f_match['name'] or '')
            fecha_final = f_match['delivery_date'] or fecha_default
            factura_final = f_match['invoice_no'] or factura_default
            qty_cmp_final = f_match['purchased_qty'] if f_match['purchased_qty'] > 0 else qty_def
            precio_final = precio_queja if (pd.notna(precio_queja) and float(precio_queja) > 0) else f_match['unit_price']
            fuente = '✅ Factura (SKU)'

        # ── NIVEL 2A: Match Catálogo por EAN
        elif ean and ean in cat_ean:
            c_match = cat_ean[ean]
            sku_final = c_match['sku'] or ref_raw
            nombre_final = c_match['detail'] or c_match['name']
            fecha_final = fecha_default
            factura_final = factura_default
            qty_cmp_final = qty_def
            precio_final = precio_queja
            fuente = '📚 Catálogo (EAN)'
            no_encontrados += 1

        # ── NIVEL 2B: Match Catálogo por SKU
        elif sku_norm_q and sku_norm_q in cat_sku:
            c_match = cat_sku[sku_norm_q]
            sku_final = c_match['sku'] or ref_raw
            nombre_final = c_match['detail'] or c_match['name']
            fecha_final = fecha_default
            factura_final = factura_default
            qty_cmp_final = qty_def
            precio_final = precio_queja
            fuente = '📚 Catálogo (SKU)'
            no_encontrados += 1


        # ── NIVEL 3: Sin datos
        else:
            sku_final = ref_raw
            nombre_final = '[Falta Factura y Catálogo]'
            fecha_final = fecha_default
            factura_final = factura_default
            qty_cmp_final = qty_def
            precio_final = precio_queja
            fuente = '❌ Sin datos'
            no_encontrados += 1


        resultados.append({
            'CIKKSZÁM / ITEM NUMBER':               sku_final,
            'MEGNEVEZÉS / NAME':                    nombre_final,
            'KISZÁLLÍTÁS DÁTUMA / DATE OF DELIVERY': fecha_final,
            'VÁSÁROLT DB / PURCHASED PCS':          qty_cmp_final,
            'ÉRINTETT MENNYISÉG / QTY CONCERNED':   qty_def,
            'NETTÓ ÁR/DB':                          precio_final,
            'INVOICE NO.':                          factura_final,
            'PANASZ OKA / REASON FOR COMPLAINT':    reason,
            '_fuente':                              fuente,
        })

    return pd.DataFrame(resultados), no_encontrados

# ─────────────────────────────────────────────
#  RELLENO Y FORMATEO DE LA PLANTILLA OFICIAL
# ─────────────────────────────────────────────

def _safe_write(ws, row: int, col: int, value):
    """Escribe un valor en una celda, resolviendo celdas combinadas si es necesario."""
    import openpyxl.cell.cell as _opc
    cell = ws.cell(row=row, column=col)
    if isinstance(cell, _opc.MergedCell):
        for merged_range in ws.merged_cells.ranges:
            if (merged_range.min_row <= row <= merged_range.max_row and
                    merged_range.min_col <= col <= merged_range.max_col):
                ws.cell(row=merged_range.min_row, column=merged_range.min_col).value = value
                return
        return
    cell.value = value

def _copy_row_style(ws, src_row: int, dst_row: int, max_col: int = 20):
    """Clona estilos, fuentes, bordes y formatos celda por celda."""
    for col in range(1, max_col + 1):
        src_cell = ws.cell(row=src_row, column=col)
        dst_cell = ws.cell(row=dst_row, column=col)
        if src_cell.has_style:
            dst_cell.font = copy(src_cell.font)
            dst_cell.border = copy(src_cell.border)
            dst_cell.fill = copy(src_cell.fill)
            dst_cell.number_format = copy(src_cell.number_format)
            dst_cell.protection = copy(src_cell.protection)
            dst_cell.alignment = copy(src_cell.alignment)

def rellenar_plantilla(wb, df_final, partner_name='Zalando', contact_name='Rodrigo', partner_request='Pay penalty', moneda='€') -> tuple:
    """
    Rellena la plantilla oficial de Home Impex al 100% de los estándares requeridos:
    - Cabecera superior (Filas 5-6): Partner, Sales assistant, Fecha
    - Columnas oficiales: Foto ('attached'), Fecha, SKU, Nombre, Motivo, Qty Comprada,
      Qty Afectada, Fórmulas de % y Total Neto, Precio, Promoción (0), Divisa ('€'), Solicitud, Factura.
    - Fila TOTAL con suma dinámica de todas las filas.
    """
    ws = wb.active

    # 1. Rellenar cabecera superior si existen los campos
    today_str = datetime.now().strftime('%d.%m.%Y')
    _safe_write(ws, 5, 3, today_str)      # Fecha llegada asistente
    _safe_write(ws, 5, 7, partner_name)   # Partner name
    _safe_write(ws, 6, 3, today_str)      # Fecha llegada quality
    _safe_write(ws, 6, 7, contact_name)   # Asistente / Sales Manager

    # 2. Localizar la fila de cabecera de la tabla (Row 11 típicamente)
    KEYWORDS_HEADER_ROW = ['CIKKSZÁM', 'ITEM NUMBER', 'MEGNEVEZÉS', 'PRODUCT NAME', 'PANASZ OKA']
    header_row_idx = -1
    for row_idx, row in enumerate(ws.iter_rows(max_row=30, values_only=True), start=1):
        row_str = ' '.join(str(v).upper() for v in row if pd.notna(v))
        if sum(1 for kw in KEYWORDS_HEADER_ROW if kw in row_str) >= 2:
            header_row_idx = row_idx
            break

    if header_row_idx == -1:
        return None, "No se encontró la cabecera de la tabla en la plantilla"

    start_row = header_row_idx + 1

    # 3. Localizar la fila original de TOTAL:
    total_row_orig = -1
    for r in range(start_row, start_row + 100):
        row_str = ' '.join(str(ws.cell(row=r, column=c).value).upper() for c in range(1, ws.max_column + 1) if ws.cell(row=r, column=c).value)
        if 'TOTAL' in row_str:
            total_row_orig = r
            break
    if total_row_orig == -1:
        total_row_orig = start_row + 10

    initial_data_rows = total_row_orig - start_row
    n_items = len(df_final)

    # 4. Ajuste dinámico del número de filas
    if n_items > initial_data_rows:
        ws.insert_rows(total_row_orig, amount=n_items - initial_data_rows)
        for r in range(start_row + initial_data_rows, start_row + n_items):
            _copy_row_style(ws, start_row, r, ws.max_column)
    elif n_items < initial_data_rows:
        ws.delete_rows(start_row + n_items, amount=initial_data_rows - n_items)

    total_row_idx = start_row + n_items
    last_data_row = start_row + n_items - 1

    # 5. Escribir datos celda por celda según el estándar oficial
    for i, (_, row_data) in enumerate(df_final.iterrows()):
        tr = start_row + i

        sku_val     = row_data.get('CIKKSZÁM / ITEM NUMBER', '')
        name_val    = row_data.get('MEGNEVEZÉS / NAME', '')
        date_val    = row_data.get('KISZÁLLÍTÁS DÁTUMA / DATE OF DELIVERY', '')
        qty_cmp_val = int(pd.to_numeric(row_data.get('VÁSÁROLT DB / PURCHASED PCS', 1), errors='coerce') or 1)
        qty_def_val = int(pd.to_numeric(row_data.get('ÉRINTETT MENNYISÉG / QTY CONCERNED', 1), errors='coerce') or 1)
        prc_raw     = row_data.get('NETTÓ ÁR/DB', None)
        prc_val     = float(prc_raw) if pd.notna(prc_raw) and str(prc_raw).strip() != '' else ''
        inv_val     = str(row_data.get('INVOICE NO.', '')).strip()
        reason_val  = str(row_data.get('PANASZ OKA / REASON FOR COMPLAINT', 'WORKMANSHIP'))

        # Col B (2): FOTÓ / PHOTO
        _safe_write(ws, tr, 2, 'attached')

        # Col C (3): KISZÁLLÍTÁS DÁTUMA
        _safe_write(ws, tr, 3, date_val)

        # Col D (4): CIKKSZÁM / ITEM NUMBER
        _safe_write(ws, tr, 4, sku_val)

        # Col F (6): MEGNEVEZÉS / NAME
        _safe_write(ws, tr, 6, name_val)

        # Col G (7): PANASZ OKA / REASON FOR COMPLAINT
        _safe_write(ws, tr, 7, reason_val)

        # Col H (8): VÁSÁROLT DB / PURCHASED PCS
        _safe_write(ws, tr, 8, qty_cmp_val)

        # Col I (9): ÉRINTETT MENNYISÉG / QUANTITY CONCERNED
        _safe_write(ws, tr, 9, qty_def_val)

        # Col J (10): VISSZÁRU MÉRTÉKE % (Fórmula =I/H)
        _safe_write(ws, tr, 10, f'=I{tr}/H{tr}')

        # Col L (12): NETTÓ ÁR/DB
        _safe_write(ws, tr, 12, prc_val)

        # Col M (13): PROMÓCIÓ % (0)
        _safe_write(ws, tr, 13, 0)

        # Col N (14): PÉNZNEM / CURRENCY ('€')
        _safe_write(ws, tr, 14, moneda)

        # Col O (15): PARTNER IGÉNYE ('Pay penalty')
        _safe_write(ws, tr, 15, partner_request)

        # Col P (16): TOTAL NETTÓ (Fórmula =(I*L)*(100%-M))
        _safe_write(ws, tr, 16, f'=(I{tr}*L{tr})*(100%-M{tr})')

        # Col Q (17): SZÁMLASZÁM / INVOICE NUMBER
        _safe_write(ws, tr, 17, inv_val)

    # 6. Actualizar la fórmula global del TOTAL:
    _safe_write(ws, total_row_idx, 16, f'=SUM(P{start_row}:P{last_data_row})')

    return wb, None

# ─────────────────────────────────────────────
#  GENERACIÓN DE EXCEL ESTÁNDAR CONSOLIDADO
# ─────────────────────────────────────────────

def generar_excel_estandar(df_export, partner_request='Pay penalty', moneda='€') -> io.BytesIO:
    output = io.BytesIO()
    df_out = df_export.copy()
    
    # Asegurar formato europeo de columnas
    if 'PHOTO' not in df_out.columns:
        df_out.insert(0, 'PHOTO', 'attached')
    if 'PROMÓCIÓ %' not in df_out.columns:
        df_out['PROMÓCIÓ %'] = 0
    if 'CURRENCY' not in df_out.columns:
        df_out['CURRENCY'] = moneda
    if 'PARTNER REQUEST' not in df_out.columns:
        df_out['PARTNER REQUEST'] = partner_request

    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df_out.to_excel(writer, index=False, sheet_name='Consolidado')
        ws = writer.sheets['Consolidado']
        header_fill = PatternFill("solid", fgColor="0F2044")
        header_font = Font(bold=True, color="FFFFFF")
        thin = Side(style='thin', color="CCCCCC")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        for col_idx, col_name in enumerate(df_out.columns, start=1):
            cell = ws.cell(row=1, column=col_idx)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', wrap_text=True)
            try:
                max_len = max(df_out[col_name].astype(str).map(len).max(), len(col_name)) + 2
            except Exception:
                max_len = 20
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len, 45)

        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=ws.max_column):
            for cell in row:
                cell.border = border

    output.seek(0)
    return output

# ─────────────────────────────────────────────
#  STREAMLIT UI — SIDEBAR
# ─────────────────────────────────────────────

st.set_page_config(
    page_title="Gestor B2B — Home Impex",
    layout="wide", page_icon="📦",
    initial_sidebar_state="expanded",
)

st.markdown("""
<style>
[data-testid="stSidebar"] { background: #0F2044; }
[data-testid="stSidebar"] * { color: #E5ECF6 !important; }
[data-testid="stSidebar"] .stTextInput input,
[data-testid="stSidebar"] .stSelectbox div[data-baseweb],
[data-testid="stSidebar"] .stNumberInput input {
    background: #1A3360 !important; border: 1px solid #2E4D80 !important;
    color: #E5ECF6 !important;
}
h1 { color: #0F2044; font-size: 2rem; font-weight: 800; }
h2, h3 { color: #1E3A8A; }
.metric-box { background:#F0F4FF; border-radius:10px; padding:16px; text-align:center; border: 1px solid #DBEAFE; }
.metric-num  { font-size:2.2rem; font-weight:800; color:#1E3A8A; }
.metric-num.warn { color:#DC2626; }
.metric-label{ font-size:0.8rem; color:#6B7280; margin-top:4px; }
.catalog-badge { background:#064E3B; color:#D1FAE5 !important; border-radius:8px;
                 padding:8px 12px; font-size:0.8rem; margin-top:8px; }
div[data-testid="stFileUploaderDropzone"] {
    border: 2px dashed #3B82F6 !important;
    border-radius: 10px !important; background: #EFF6FF !important;
}
</style>
""", unsafe_allow_html=True)

for k, v in [('df_final', None), ('log_msgs', []), ('catalogo', None), ('catalogo_info', '')]:
    if k not in st.session_state:
        st.session_state[k] = v

with st.sidebar:
    st.markdown("## 🏢 Datos del Caso")
    st.divider()
    partner_name    = st.text_input("Cliente / Partner", value="Zalando")
    contact_name    = st.text_input("Contacto / Asistente", value="Rodrigo")
    ref_caso        = st.text_input("Referencia del caso / DNZ", placeholder="ej. DNZ-2026-016699")
    partner_request = st.selectbox("Solicitud del Partner", REQUEST_OPCIONES, index=0)
    moneda          = st.selectbox("Divisa / Moneda", ["€", "EUR", "HUF", "GBP", "USD", "PLN"], index=0)
    fecha_manual    = st.text_input(
        "📅 Fecha de Entrega (override opcional)",
        placeholder="ej. 12.06.2026",
        help="Si se deja vacío, se extrae automáticamente de las facturas históricas/PDFs"
    )
    factura_manual  = st.text_input(
        "📄 Nº Factura (override opcional)",
        placeholder="ej. D26/00720",
        help="Si se deja vacío, se extrae automáticamente de las facturas históricas/PDFs"
    )

    st.divider()
    st.markdown("## 🏷️ Motivo por defecto")
    reason_default = st.selectbox("Si no hay motivo en el archivo", REASON_OPCIONES)

    st.divider()
    st.markdown("## 📚 Catálogo Maestro")
    catalogo_override = st.file_uploader(
        "Override manual (opcional)", type=["xlsx", "xls"],
        help="Si subes aquí un catálogo, tiene prioridad sobre la carpeta /catalogo/"
    )

    cat_bytes = None
    cat_name  = ''
    if catalogo_override:
        cat_bytes = catalogo_override.read()
        cat_name  = catalogo_override.name

    catalogo_dict, catalogo_info = cargar_catalogo_maestro(
        CARPETA_CATALOGO, cat_bytes, cat_name
    )
    st.markdown(f'<div class="catalog-badge">{catalogo_info}</div>', unsafe_allow_html=True)
    st.divider()
    st.caption("📦 Home Impex — Gestor Automatizado B2B v3.0")

# ─────────────────────────────────────────────
#  MAIN LAYOUT & PROCESAMIENTO
# ─────────────────────────────────────────────

st.title("📦 Gestor Automatizado de Reclamaciones B2B")
st.markdown("Cruza automáticamente reportes de incidencias con facturación histórica y catálogo maestro.")
st.divider()

col_templ, col_data = st.columns([1, 2], gap="large")

with col_templ:
    st.markdown("#### 📄 Plantilla Oficial *(opcional)*")
    st.caption("Si la subes, el resultado se inyectará directamente en ella con diseño oficial.")
    plantilla_file = st.file_uploader(
        "Plantilla vacía", type=["xlsx"],
        label_visibility="collapsed", key="plantilla"
    )
    if plantilla_file:
        st.success(f"✅ `{plantilla_file.name}`")

with col_data:
    st.markdown("#### 📂 Reportes de Quejas + Facturas / POs / Históricos")
    st.caption("Arrastra Excel, CSV y/o PDFs de facturas. El motor clasifica y cruza automáticamente.")
    uploaded_files = st.file_uploader(
        "Archivos de datos", type=["xlsx", "xls", "csv", "pdf"],
        accept_multiple_files=True,
        label_visibility="collapsed", key="datos"
    )

st.divider()
btn_col, _ = st.columns([1, 3])
process = btn_col.button(
    "🚀  Procesar y Cruzar Datos", type="primary",
    disabled=not uploaded_files, use_container_width=True
)

if process and uploaded_files:
    st.session_state['df_final'] = None
    df_quejas_master   = pd.DataFrame()
    df_facturas_master = pd.DataFrame()
    log = []

    with st.status("🔍 Analizando y clasificando documentos...", expanded=True) as status:
        for uf in uploaded_files:
            df, tipo = leer_archivo(uf)
            icon_tipo = {
                'ignorar':       ('⚠️', 'Ignorado (plantilla)'),
                'sin_ean':       ('⚠️', 'Sin columnas identificables'),
                'pdf_sin_datos': ('⚠️', 'PDF sin artículos extraíbles (o Nota de Crédito)'),
                'factura':       ('🟢', '**Factura / Base de datos**'),
                'queja':         ('📋', '**Reporte de quejas / RSC**'),
            }
            if isinstance(tipo, str) and tipo.startswith('error'):
                st.error(f"❌ `{uf.name}` — {tipo}")
                continue
            icon, label = icon_tipo.get(tipo, ('❓', 'No clasificado'))
            rows = len(df) if df is not None else 0
            msg = f"{icon} `{uf.name}` → {label} · *{rows} filas*"
            st.write(msg)
            log.append(msg)

            if tipo == 'factura' and df is not None:
                df_facturas_master = pd.concat([df_facturas_master, df], ignore_index=True)
            elif tipo == 'queja' and df is not None:
                df_quejas_master = pd.concat([df_quejas_master, df], ignore_index=True)

        if df_quejas_master.empty and df_facturas_master.empty:
            status.update(label="❌ Sin archivos procesables.", state="error")
            st.stop()
        st.write(f"---\n📊 **{len(df_quejas_master)}** artículos en quejas · **{len(df_facturas_master)}** líneas en facturas")
        status.update(label="✅ Clasificación completada.", state="complete", expanded=False)

    if df_quejas_master.empty:
        st.warning("No se encontraron reportes de quejas en los archivos subidos.")
        st.stop()

    with st.spinner("Ejecutando cruce en cascada multi-nivel..."):
        df_final, no_enc = cruzar_datos(
            df_quejas_master, df_facturas_master, catalogo_dict, reason_default,
            fecha_manual_override=fecha_manual, factura_manual_override=factura_manual,
            partner_request=partner_request, moneda=moneda
        )
    st.session_state['df_final'] = df_final

# ─────────────────────────────────────────────
#  VISTA DE RESULTADOS Y DESCARGAS
# ─────────────────────────────────────────────

if st.session_state['df_final'] is not None:
    df_final = st.session_state['df_final']
    no_enc   = (df_final['INVOICE NO.'].isin(['(not found)', '(no factura periodo)', ''])).sum()
    total_eur = (
        df_final['NETTÓ ÁR/DB'].fillna(0) *
        df_final['ÉRINTETT MENNYISÉG / QTY CONCERNED'].fillna(0)
    ).sum()

    fuentes = df_final['_fuente'].value_counts().to_dict()

    st.subheader("📊 Resultados del Cruce Consolidado")
    c1, c2, c3, c4 = st.columns(4)
    c1.markdown(f'<div class="metric-box"><div class="metric-num">{len(df_final)}</div><div class="metric-label">Artículos Procesados</div></div>', unsafe_allow_html=True)
    c2.markdown(f'<div class="metric-box"><div class="metric-num {"warn" if no_enc>0 else ""}">{no_enc}</div><div class="metric-label">Sin Factura Específica</div></div>', unsafe_allow_html=True)
    c3.markdown(f'<div class="metric-box"><div class="metric-num">{df_final["INVOICE NO."].replace("", "(vacío)").nunique()}</div><div class="metric-label">Facturas Vinculadas</div></div>', unsafe_allow_html=True)
    c4.markdown(f'<div class="metric-box"><div class="metric-num">{total_eur:,.2f} {moneda}</div><div class="metric-label">Importe Total Estimado</div></div>', unsafe_allow_html=True)

    fuente_str = "  ·  ".join(f"{k}: **{v}**" for k, v in fuentes.items())
    st.caption(f"Origen de datos: {fuente_str}")
    st.markdown("---")

    df_edit = df_final.drop(columns=['_fuente'], errors='ignore')
    st.markdown("#### ✏️ Revisa y edita los datos antes de exportar")

    df_editado = st.data_editor(
        df_edit,
        use_container_width=True,
        num_rows="dynamic",
        height=320,
        column_config={
            'PANASZ OKA / REASON FOR COMPLAINT': st.column_config.SelectboxColumn(
                "Motivo de Queja", options=REASON_OPCIONES, required=False,
            ),
            'KISZÁLLÍTÁS DÁTUMA / DATE OF DELIVERY': st.column_config.TextColumn(
                "Fecha de Entrega", help="Extraída de la factura. Editable."
            ),
            'INVOICE NO.': st.column_config.TextColumn(
                "Nº Factura", help="Número de factura o albarán."
            ),
            'NETTÓ ÁR/DB': st.column_config.NumberColumn("Precio Neto/UD", format="%.2f"),
            'VÁSÁROLT DB / PURCHASED PCS':        st.column_config.NumberColumn("Uds Compradas", step=1),
            'ÉRINTETT MENNYISÉG / QTY CONCERNED': st.column_config.NumberColumn("Uds Afectadas", step=1),
        },
    )


    st.markdown("---")
    st.markdown("#### 💾 Descargar Formularios Generados")
    dl1, dl2 = st.columns(2)

    # 1. Descarga plantilla oficial si fue subida
    if plantilla_file:
        try:
            plantilla_file.seek(0)
            wb_base = openpyxl.load_workbook(io.BytesIO(plantilla_file.read()))
            wb_relleno, err = rellenar_plantilla(
                wb_base, df_editado,
                partner_name=partner_name, contact_name=contact_name,
                partner_request=partner_request, moneda=moneda
            )
            if err:
                dl1.error(f"Error plantilla: {err}")
            else:
                buf = io.BytesIO()
                wb_relleno.save(buf)
                buf.seek(0)
                nombre_dl = f"LISTO_{partner_name}__{plantilla_file.name}"
                dl1.download_button(
                    label=f"📄 Descargar Plantilla Oficial Rellena",
                    data=buf,
                    file_name=nombre_dl,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True,
                    type="primary"
                )
        except Exception as e:
            dl1.error(f"Error al rellenar plantilla: {e}")
    else:
        dl1.info("ℹ️ Sube la plantilla oficial a la izquierda para generar el formato oficial de Home Impex.")

    # 2. Descarga Excel Consolidado siempre disponible
    buf_std = generar_excel_estandar(df_editado, partner_request=partner_request, moneda=moneda)
    dl2.download_button(
        label="📊 Descargar Excel Consolidado Estándar",
        data=buf_std,
        file_name=f"LISTO_{partner_name}__Consolidado.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True
    )
